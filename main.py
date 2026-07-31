"""
main.py
-------
The ENTIRE backend in one file, mirroring the pattern used in the Finance
Management System (GitHub-as-database, no SQL database, Gemini does the
natural-language classification, openpyxl does the bookkeeping).

WHAT THIS SYSTEM IS
  An Inventory Management System with an "AI Warehouse Assistant" chat bot.
  You type things like:
    "received 200 units of Dell Laptop in Mumbai Warehouse"
    "dispatched 50 Dell Laptop from Mumbai Warehouse to Client XYZ"
    "moved 30 Dell Laptop from Mumbai Warehouse to Delhi Warehouse"
    "lost 5 Dell Laptop in Delhi Warehouse due to damage"
    "how much Dell Laptop do we have and where is it kept?"
    "what's been sitting in stock for too long?"
  Gemini classifies the message, the backend applies it to a FIFO batch
  ledger (or, for questions, computes the answer deterministically from
  the ledger - Gemini only extracts WHAT is being asked, it never
  invents the numbers), and everything is saved back to an .xlsx file
  in a GitHub repo, downloadable at any time.

KEY DESIGN DECISION - UNLIKE THE FINANCE SYSTEM, THE CATALOG IS OPEN:
  The finance system had a fixed master template (fixed line items) that
  was never restructured, only its values changed. Here products AND
  locations are NOT fixed - the assistant can create a brand-new product
  or location the first time it's mentioned. To avoid duplicate near-
  identical products ("Dell Laptop" vs "Dell laptops"), every prompt to
  Gemini includes the current distinct list of known products/locations
  and instructs it to reuse an existing exact name whenever the message
  is clearly referring to the same thing.

FIFO MODEL:
  Stock is never stored as a single "quantity per product" number. It is
  stored as individual BATCHES (Product, Location, Date Received, Qty
  Received, Qty Remaining). Every stock-out (dispatch/adjustment-down/
  transfer-out) consumes the OLDEST active batches first for that exact
  (Product, Location) pair. This is what lets the system answer "what's
  been sitting around too long" - it's a direct query over batch ages,
  not an estimate.

Environment variables needed (set these in Vercel Project Settings):
  GITHUB_TOKEN    - GitHub Personal Access Token (Contents: Read & Write)
  GITHUB_REPO     - "your-username/your-repo"
  GITHUB_BRANCH   - usually "main"
  EXCEL_PATH      - where the live .xlsx lives in the repo, e.g. "data/inventory_data.xlsx"
  ACCESS_CODE     - the shared password used to log in
  GEMINI_API_KEY  - Google Gemini API key (from .env / platform secrets - never hardcoded)
                    (model is fixed to gemini-3.1-flash-lite, not configurable via env)
  AGING_THRESHOLD_DAYS - optional, defaults to 90 (roughly 3 months)
"""

import os
import io
import re
import json
import base64
from datetime import datetime, timezone, date
from typing import Dict, List, Optional

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel


# ============================================================================
# SETTINGS
# ============================================================================
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO = os.environ.get("GITHUB_REPO", "")
GITHUB_BRANCH = os.environ.get("GITHUB_BRANCH", "main")
EXCEL_PATH = os.environ.get("EXCEL_PATH", "data/inventory_data.xlsx")
ACCESS_CODE = os.environ.get("ACCESS_CODE", "inventory2026")

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL = "gemini-3.1-flash-lite"  # fixed per requirements - do not swap models
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

AGING_THRESHOLD_DAYS = int(os.environ.get("AGING_THRESHOLD_DAYS", "90"))

GITHUB_API = "https://api.github.com"
GH_HEADERS = {
    "Authorization": f"Bearer {GITHUB_TOKEN}",
    "Accept": "application/vnd.github+json",
}

# ----------------------------------------------------------------------------
# Sheet layout. No master template file is needed (unlike the finance
# system) because there are no fixed line items here - a brand-new
# workbook with just headers is the correct starting point.
# ----------------------------------------------------------------------------
BATCH_SHEET = "Batches"
BATCH_HEADERS = ["Batch ID", "Product", "Location", "Date Received",
                  "Qty Received", "Qty Remaining", "Unit", "Status"]

LOG_SHEET = "Transaction Log"
LOG_HEADERS = ["Timestamp", "Original Message", "Action", "Product",
               "From Location", "To Location", "Quantity", "Unit",
               "Entered By", "Status", "Reversal Data"]

ACTIONS = {"RECEIVE", "DISPATCH", "TRANSFER", "ADJUST"}


# ============================================================================
# GITHUB READ / WRITE (the "database") - identical pattern to the finance system
# ============================================================================
def _contents_url(path):
    return f"{GITHUB_API}/repos/{GITHUB_REPO}/contents/{path}"


def _get_file_meta():
    """Returns (sha, raw_bytes). raw_bytes is None if the file doesn't exist yet."""
    if not GITHUB_TOKEN or not GITHUB_REPO:
        raise RuntimeError("GITHUB_TOKEN / GITHUB_REPO are not set on the server.")
    resp = requests.get(_contents_url(EXCEL_PATH), headers=GH_HEADERS, params={"ref": GITHUB_BRANCH}, timeout=30)
    if resp.status_code == 404:
        return None, None
    resp.raise_for_status()
    data = resp.json()
    return data["sha"], base64.b64decode(data["content"])


def _put_file(raw_bytes, commit_message):
    sha, _ = _get_file_meta()
    payload = {
        "message": commit_message,
        "content": base64.b64encode(raw_bytes).decode(),
        "branch": GITHUB_BRANCH,
    }
    if sha:
        payload["sha"] = sha
    resp = requests.put(_contents_url(EXCEL_PATH), headers=GH_HEADERS, json=payload, timeout=30)
    resp.raise_for_status()
    return resp.json()


def get_download_url():
    return f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{EXCEL_PATH}"


# ============================================================================
# EXCEL SERVICE
# ============================================================================
_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _style_header(ws: Worksheet, headers: List[str]):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


class ExcelService:

    @classmethod
    def new_workbook(cls) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = BATCH_SHEET
        ws.append(BATCH_HEADERS)
        _style_header(ws, BATCH_HEADERS)
        log = wb.create_sheet(LOG_SHEET)
        log.append(LOG_HEADERS)
        _style_header(log, LOG_HEADERS)
        return wb

    @classmethod
    def load_live_workbook(cls) -> Workbook:
        _, raw = _get_file_meta()
        if raw is None:
            return cls.new_workbook()
        wb = load_workbook(io.BytesIO(raw), data_only=False)
        if BATCH_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(BATCH_SHEET)
            ws.append(BATCH_HEADERS)
            _style_header(ws, BATCH_HEADERS)
        if LOG_SHEET not in wb.sheetnames:
            log = wb.create_sheet(LOG_SHEET)
            log.append(LOG_HEADERS)
            _style_header(log, LOG_HEADERS)
        return wb

    @classmethod
    def save(cls, wb: Workbook, message: str):
        buf = io.BytesIO()
        wb.save(buf)
        return _put_file(buf.getvalue(), message)

    # -- Batches -------------------------------------------------------------
    @classmethod
    def _batch_ws(cls, wb: Workbook) -> Worksheet:
        return wb[BATCH_SHEET]

    @classmethod
    def next_batch_id(cls, wb: Workbook) -> int:
        ws = cls._batch_ws(wb)
        max_id = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] is not None:
                max_id = max(max_id, int(row[0]))
        return max_id + 1

    @classmethod
    def all_batches(cls, wb: Workbook) -> List[dict]:
        ws = cls._batch_ws(wb)
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            out.append(dict(zip(BATCH_HEADERS, row)))
        return out

    @classmethod
    def known_products(cls, wb: Workbook) -> List[str]:
        return sorted({b["Product"] for b in cls.all_batches(wb) if b["Product"]})

    @classmethod
    def known_locations(cls, wb: Workbook) -> List[str]:
        return sorted({b["Location"] for b in cls.all_batches(wb) if b["Location"]})

    @classmethod
    def add_batch(cls, wb: Workbook, product: str, location: str, date_received: str,
                  qty: float, unit: str) -> int:
        ws = cls._batch_ws(wb)
        batch_id = cls.next_batch_id(wb)
        ws.append([batch_id, product, location, date_received, qty, qty, unit, "Active"])
        return batch_id

    @classmethod
    def _batch_rows(cls, wb: Workbook):
        """Yields (row_cells) for every batch row, cells indexable like BATCH_HEADERS."""
        ws = cls._batch_ws(wb)
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            yield row

    @classmethod
    def active_batches_fifo(cls, wb: Workbook, product: str, location: str):
        """Active batch ROWS (cell objects) for an exact product+location,
        oldest Date Received first. Matching is case-insensitive/trimmed."""
        rows = []
        for row in cls._batch_rows(wb):
            if (str(row[1].value).strip().lower() == product.strip().lower()
                    and str(row[2].value).strip().lower() == location.strip().lower()
                    and (row[5].value or 0) > 0):
                rows.append(row)
        rows.sort(key=lambda r: str(r[3].value))
        return rows

    @classmethod
    def consume_fifo(cls, wb: Workbook, product: str, location: str, qty: float):
        """Deducts qty from the oldest active batches at (product, location).
        Returns (consumed_breakdown, shortfall) where consumed_breakdown is a
        list of {batch_id, date_received, qty_taken} and shortfall is how much
        could NOT be fulfilled (0 if fully satisfied). Caller decides whether
        a shortfall is acceptable."""
        remaining_needed = qty
        breakdown = []
        for row in cls.active_batches_fifo(wb, product, location):
            if remaining_needed <= 0:
                break
            available = row[5].value or 0
            take = min(available, remaining_needed)
            if take <= 0:
                continue
            row[5].value = round(available - take, 4)
            row[7].value = "Active" if row[5].value > 0 else "Depleted"
            breakdown.append({"batch_id": row[0].value, "date_received": str(row[3].value),
                               "qty_taken": take})
            remaining_needed -= take
        shortfall = round(max(remaining_needed, 0), 4)
        return breakdown, shortfall

    @classmethod
    def restore_fifo(cls, wb: Workbook, breakdown: List[dict]):
        """Exact reversal of consume_fifo, used by undo - adds each taken
        amount back to its exact original batch by Batch ID (never guesses)."""
        rows_by_id = {row[0].value: row for row in cls._batch_rows(wb)}
        for item in breakdown:
            row = rows_by_id.get(item["batch_id"])
            if row is None:
                continue
            row[5].value = round((row[5].value or 0) + item["qty_taken"], 4)
            row[7].value = "Active"

    @classmethod
    def stock_summary(cls, wb: Workbook, product: Optional[str] = None):
        """{'Product Name': {'Location A': qty, 'Location B': qty, ...}, ...}
        restricted to Active/remaining>0 batches. Case preserved from the
        first-seen spelling; matching for filtering is case-insensitive."""
        out: Dict[str, Dict[str, float]] = {}
        for b in cls.all_batches(wb):
            if (b["Qty Remaining"] or 0) <= 0:
                continue
            if product and product.strip().lower() != str(b["Product"]).strip().lower():
                continue
            out.setdefault(b["Product"], {})
            out[b["Product"]][b["Location"]] = round(out[b["Product"]].get(b["Location"], 0) + b["Qty Remaining"], 4)
        return out

    @classmethod
    def aging_batches(cls, wb: Workbook, threshold_days: int = AGING_THRESHOLD_DAYS):
        today = date.today()
        out = []
        for b in cls.all_batches(wb):
            if (b["Qty Remaining"] or 0) <= 0:
                continue
            try:
                d = _parse_date(b["Date Received"])
            except Exception:
                continue
            age = (today - d).days
            if age >= threshold_days:
                out.append({**b, "Age Days": age})
        return sorted(out, key=lambda r: -r["Age Days"])

    # -- Transaction Log -------------------------------------------------------
    @classmethod
    def append_log(cls, wb: Workbook, original_msg: str, action: str, product: str,
                   from_location: Optional[str], to_location: Optional[str],
                   qty: Optional[float], unit: Optional[str], entered_by: str,
                   status: str, reversal_data: Optional[list] = None) -> int:
        ws = wb[LOG_SHEET]
        now = datetime.now(timezone.utc).isoformat()
        ws.append([now, original_msg, action, product, from_location, to_location,
                   qty, unit, entered_by, status, json.dumps(reversal_data or [])])
        return ws.max_row

    @classmethod
    def get_log_rows(cls, wb: Workbook, limit: Optional[int] = None):
        ws = wb[LOG_SHEET]
        rows = [dict(zip(LOG_HEADERS, r)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0] is not None]
        rows.reverse()  # newest first
        return rows[:limit] if limit else rows

    @classmethod
    def undo_last(cls, wb: Workbook):
        ws = wb[LOG_SHEET]
        target_row = None
        for r in range(ws.max_row, 1, -1):
            if ws.cell(row=r, column=10).value == "Applied":
                target_row = r
                break
        if target_row is None:
            raise ValueError("No applied transaction to undo.")

        action = ws.cell(row=target_row, column=3).value
        product = ws.cell(row=target_row, column=4).value
        from_loc = ws.cell(row=target_row, column=5).value
        to_loc = ws.cell(row=target_row, column=6).value
        qty = ws.cell(row=target_row, column=7).value
        unit = ws.cell(row=target_row, column=8).value
        reversal_data = json.loads(ws.cell(row=target_row, column=11).value or "[]")

        if action == "RECEIVE":
            # reversal_data holds the single batch_id that was created
            for item in reversal_data:
                for row in cls._batch_rows(wb):
                    if row[0].value == item["batch_id"]:
                        row[5].value = 0
                        row[7].value = "Depleted"
        elif action == "DISPATCH":
            cls.restore_fifo(wb, reversal_data)
        elif action == "ADJUST":
            if qty and qty > 0:
                # was a positive adjustment (new batch) -> deplete it
                for item in reversal_data:
                    for row in cls._batch_rows(wb):
                        if row[0].value == item["batch_id"]:
                            row[5].value = 0
                            row[7].value = "Depleted"
            else:
                cls.restore_fifo(wb, reversal_data)
        elif action == "TRANSFER":
            # reversal_data = {"consumed": [...], "created_batch_ids": [...]}
            cls.restore_fifo(wb, reversal_data.get("consumed", []))
            for bid in reversal_data.get("created_batch_ids", []):
                for row in cls._batch_rows(wb):
                    if row[0].value == bid:
                        row[5].value = 0
                        row[7].value = "Depleted"

        ws.cell(row=target_row, column=10, value="Reversed")
        cls.append_log(wb, f"UNDO: {ws.cell(row=target_row, column=2).value}", action, product,
                        from_loc, to_loc, qty, unit, "system", "Applied (undo)")
        return {"action": action, "product": product, "quantity": qty}


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


# ============================================================================
# GEMINI SERVICE
# Classifies a free-text message into either a STOCK MOVEMENT or a QUERY.
# Gemini never computes quantities itself - it only extracts structured
# intent; every number in the final answer comes from the Batches sheet.
# ============================================================================
class GeminiService:

    SYSTEM_TEMPLATE = """You are the parsing engine behind a warehouse inventory chat bot.
Read the user's message and return ONLY a single JSON object (no markdown
fences, no commentary) matching exactly one of these shapes:

1) Stock movement:
{{
  "type": "movement",
  "action": "RECEIVE" | "DISPATCH" | "TRANSFER" | "ADJUST",
  "product": "<name>",
  "location": "<name>",            // required for RECEIVE, DISPATCH, ADJUST
  "from_location": "<name>",       // required for TRANSFER only
  "to_location": "<name>",         // required for TRANSFER only
  "quantity": <number>,             // for ADJUST, positive = found extra stock, negative = write-off/loss/damage
  "unit": "<e.g. units, boxes, kg>",
  "date_received": "YYYY-MM-DD"     // only for RECEIVE; the date stock arrived. Use today if not stated.
}}

2) Stock question ("how much do we have", "where is it kept", "what do we have in X location"):
{{
  "type": "query",
  "query_kind": "PRODUCT_LOOKUP" | "LOCATION_LOOKUP" | "AGING_REPORT",
  "product": "<name or null>",
  "location": "<name or null>"
}}

3) Could not classify:
{{
  "type": "unknown",
  "reason": "<short reason>"
}}

RULES:
- KNOWN PRODUCTS SO FAR: {known_products}
- KNOWN LOCATIONS SO FAR: {known_locations}
- If the message clearly refers to a product/location that closely matches
  one already in the known lists above (different casing, singular/plural,
  minor typo), reuse the EXACT existing spelling from the list. Only use a
  brand-new name when it is genuinely a different item/place.
- "received", "arrived", "bought", "purchased", "added to stock", "brought in"
  -> RECEIVE.
- "dispatched", "sold", "shipped out", "sent to customer", "used", "consumed",
  "issued" -> DISPATCH.
- "moved", "transferred", "shifted" between two locations -> TRANSFER.
- "damaged", "lost", "expired", "written off", "found extra", "stock count
  correction" -> ADJUST (negative quantity for loss/damage/write-off,
  positive for found/extra).
- Questions about quantity/location/"how long has X been sitting"/"what's
  old stock"/"what's been here for months" -> type "query".
- If essential info is missing (e.g. no quantity, no product) return "unknown"
  with a clear reason.
- Never invent a product or location that has no reasonable link to the
  message text.
Return ONLY the JSON object, nothing else.
"""

    @classmethod
    def parse_message(cls, message: str, known_products: List[str], known_locations: List[str]) -> dict:
        if not GEMINI_API_KEY:
            raise RuntimeError("GEMINI_API_KEY is not set on the server.")
        system_prompt = cls.SYSTEM_TEMPLATE.format(
            known_products=", ".join(known_products) or "(none yet)",
            known_locations=", ".join(known_locations) or "(none yet)",
        )
        payload = {
            "system_instruction": {"parts": [{"text": system_prompt}]},
            "contents": [{"role": "user", "parts": [{"text": message}]}],
            "generationConfig": {"temperature": 0.1, "response_mime_type": "application/json"},
        }
        resp = requests.post(f"{GEMINI_URL}?key={GEMINI_API_KEY}", json=payload, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        try:
            text = data["candidates"][0]["content"]["parts"][0]["text"]
        except (KeyError, IndexError) as e:
            raise RuntimeError(f"Unexpected Gemini response shape: {data}") from e
        text = re.sub(r"^```json|```$", "", text.strip(), flags=re.MULTILINE).strip()
        return json.loads(text)


# ============================================================================
# INVENTORY ACTION SERVICE
# Applies a parsed movement to the Batches sheet and returns a human-readable
# result. This is where FIFO logic actually gets used.
# ============================================================================
class InventoryService:

    @classmethod
    def apply_movement(cls, wb: Workbook, parsed: dict, original_msg: str, entered_by: str) -> dict:
        action = parsed.get("action")
        product = (parsed.get("product") or "").strip()
        unit = (parsed.get("unit") or "units").strip()
        qty = parsed.get("quantity")

        if action not in ACTIONS:
            raise ValueError(f"Unknown action '{action}'.")
        if not product:
            raise ValueError("No product specified.")

        if action == "RECEIVE":
            location = (parsed.get("location") or "").strip()
            if not location:
                raise ValueError("No location specified for received stock.")
            if not qty or qty <= 0:
                raise ValueError("Quantity to receive must be a positive number.")
            date_received = parsed.get("date_received") or date.today().isoformat()
            batch_id = ExcelService.add_batch(wb, product, location, date_received, qty, unit)
            ExcelService.append_log(wb, original_msg, "RECEIVE", product, None, location,
                                     qty, unit, entered_by, "Applied",
                                     [{"batch_id": batch_id}])
            return {"action": "RECEIVE", "product": product, "location": location,
                    "quantity": qty, "unit": unit, "date_received": date_received,
                    "message": f"Recorded {qty} {unit} of {product} received at {location} "
                               f"(dated {date_received})."}

        if action == "DISPATCH":
            location = (parsed.get("location") or "").strip()
            if not location:
                raise ValueError("No location specified for dispatch.")
            if not qty or qty <= 0:
                raise ValueError("Quantity to dispatch must be a positive number.")
            breakdown, shortfall = ExcelService.consume_fifo(wb, product, location, qty)
            if shortfall > 0:
                # Roll back what we just took, since a partial dispatch could
                # be wrong for the user's bookkeeping - ask instead of guessing.
                ExcelService.restore_fifo(wb, breakdown)
                available = sum(v for locs in ExcelService.stock_summary(wb, product).values()
                                 for k, v in locs.items() if k == location)
                raise ValueError(f"Only {available} {unit} of {product} available at {location}, "
                                  f"cannot dispatch {qty}.")
            ExcelService.append_log(wb, original_msg, "DISPATCH", product, location, None,
                                     qty, unit, entered_by, "Applied", breakdown)
            oldest = breakdown[0]["date_received"] if breakdown else None
            return {"action": "DISPATCH", "product": product, "location": location,
                    "quantity": qty, "unit": unit,
                    "message": f"Dispatched {qty} {unit} of {product} from {location} "
                               f"(FIFO: oldest batch used first{', from ' + oldest if oldest else ''})."}

        if action == "TRANSFER":
            from_location = (parsed.get("from_location") or "").strip()
            to_location = (parsed.get("to_location") or "").strip()
            if not from_location or not to_location:
                raise ValueError("Both from_location and to_location are required for a transfer.")
            if not qty or qty <= 0:
                raise ValueError("Quantity to transfer must be a positive number.")
            breakdown, shortfall = ExcelService.consume_fifo(wb, product, from_location, qty)
            if shortfall > 0:
                ExcelService.restore_fifo(wb, breakdown)
                raise ValueError(f"Not enough {product} at {from_location} to transfer {qty} {unit}.")
            created_ids = []
            for item in breakdown:
                bid = ExcelService.add_batch(wb, product, to_location, item["date_received"],
                                              item["qty_taken"], unit)
                created_ids.append(bid)
            ExcelService.append_log(wb, original_msg, "TRANSFER", product, from_location, to_location,
                                     qty, unit, entered_by, "Applied",
                                     {"consumed": breakdown, "created_batch_ids": created_ids})
            return {"action": "TRANSFER", "product": product, "from_location": from_location,
                    "to_location": to_location, "quantity": qty, "unit": unit,
                    "message": f"Transferred {qty} {unit} of {product} from {from_location} to "
                               f"{to_location} (original batch dates preserved for aging)."}

        if action == "ADJUST":
            location = (parsed.get("location") or "").strip()
            if not location:
                raise ValueError("No location specified for the adjustment.")
            if qty is None or qty == 0:
                raise ValueError("Adjustment quantity must be a non-zero number.")
            if qty > 0:
                batch_id = ExcelService.add_batch(wb, product, location, date.today().isoformat(), qty, unit)
                ExcelService.append_log(wb, original_msg, "ADJUST", product, None, location,
                                         qty, unit, entered_by, "Applied", [{"batch_id": batch_id}])
                return {"action": "ADJUST", "product": product, "location": location,
                        "quantity": qty, "unit": unit,
                        "message": f"Added {qty} {unit} of {product} to {location} as a stock-count correction."}
            else:
                breakdown, shortfall = ExcelService.consume_fifo(wb, product, location, abs(qty))
                if shortfall > 0:
                    ExcelService.restore_fifo(wb, breakdown)
                    raise ValueError(f"Cannot write off {abs(qty)} {unit} of {product} at {location} - "
                                      f"only {abs(qty) - shortfall} {unit} on hand.")
                ExcelService.append_log(wb, original_msg, "ADJUST", product, location, None,
                                         qty, unit, entered_by, "Applied", breakdown)
                return {"action": "ADJUST", "product": product, "location": location,
                        "quantity": qty, "unit": unit,
                        "message": f"Wrote off {abs(qty)} {unit} of {product} at {location}."}

        raise ValueError("Unhandled action.")

    @classmethod
    def answer_query(cls, wb: Workbook, parsed: dict) -> dict:
        kind = parsed.get("query_kind")
        product = parsed.get("product")
        location = parsed.get("location")

        if kind == "AGING_REPORT":
            rows = ExcelService.aging_batches(wb)
            if not rows:
                return {"type": "query", "query_kind": kind,
                        "message": f"Nothing has been sitting in stock for {AGING_THRESHOLD_DAYS}+ days. All clear.",
                        "rows": []}
            lines = [f"{r['Product']} at {r['Location']}: {r['Qty Remaining']} {r['Unit']} "
                     f"(received {r['Date Received']}, {r['Age Days']} days old)" for r in rows[:15]]
            return {"type": "query", "query_kind": kind,
                    "message": f"{len(rows)} batch(es) have been in stock {AGING_THRESHOLD_DAYS}+ days:\n" + "\n".join(lines),
                    "rows": rows}

        if kind == "LOCATION_LOOKUP" and location:
            all_summary = ExcelService.stock_summary(wb)
            found = {p: locs[location] for p, locs in all_summary.items() if location in locs}
            if not found:
                return {"type": "query", "query_kind": kind,
                        "message": f"No stock currently recorded at {location}.", "rows": []}
            lines = [f"{p}: {q}" for p, q in sorted(found.items())]
            return {"type": "query", "query_kind": kind,
                    "message": f"Stock currently at {location}:\n" + "\n".join(lines),
                    "rows": found}

        # default: PRODUCT_LOOKUP
        if not product:
            return {"type": "query", "query_kind": kind,
                    "message": "I couldn't tell which product you're asking about.", "rows": []}
        summary = ExcelService.stock_summary(wb, product)
        if not summary:
            known = ExcelService.known_products(wb)
            close = [p for p in known if product.strip().lower() in p.lower() or p.lower() in product.strip().lower()]
            hint = f" Did you mean: {', '.join(close)}?" if close else ""
            return {"type": "query", "query_kind": kind,
                    "message": f"No stock currently recorded for '{product}'.{hint}", "rows": []}
        product_name = next(iter(summary))
        locs = summary[product_name]
        total = round(sum(locs.values()), 4)
        breakdown = "; ".join(f"{loc}: {q}" for loc, q in sorted(locs.items()))
        return {"type": "query", "query_kind": kind,
                "message": f"{product_name} - {total} total, kept at: {breakdown}.",
                "rows": locs}


# ============================================================================
# API
# ============================================================================
app = FastAPI(title="Inventory Management System - AI Warehouse Assistant")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


@app.exception_handler(Exception)
async def _all_exceptions(request, exc: Exception):
    return JSONResponse(status_code=500, content={"detail": f"{type(exc).__name__}: {exc}"})


class LoginIn(BaseModel):
    name: str
    access_code: str


class ChatIn(BaseModel):
    message: str
    entered_by: str


@app.post("/api/login")
def login(data: LoginIn):
    if data.access_code != ACCESS_CODE:
        return {"success": False, "message": "Incorrect access code."}
    if not data.name.strip():
        return {"success": False, "message": "Please enter your name."}
    return {"success": True, "name": data.name.strip()}


@app.post("/api/chat")
def chat(data: ChatIn):
    if not data.message.strip():
        raise HTTPException(400, "message is empty")

    wb = ExcelService.load_live_workbook()
    known_products = ExcelService.known_products(wb)
    known_locations = ExcelService.known_locations(wb)

    try:
        parsed = GeminiService.parse_message(data.message, known_products, known_locations)
    except Exception as e:
        raise HTTPException(502, f"Gemini request failed: {e}")

    if parsed.get("type") == "unknown":
        ExcelService.append_log(wb, data.message, "REJECTED", parsed.get("product", ""), None, None,
                                 None, None, data.entered_by, "Rejected")
        try:
            ExcelService.save(wb, f"Log rejected message ({data.entered_by})")
        except Exception:
            pass
        return {"success": False, "type": "unknown",
                "message": parsed.get("reason", "Could not understand that message.")}

    if parsed.get("type") == "query":
        result = InventoryService.answer_query(wb, parsed)
        return {"success": True, **result}

    # movement
    try:
        result = InventoryService.apply_movement(wb, parsed, data.message, data.entered_by)
    except ValueError as e:
        return {"success": False, "type": "movement_rejected", "message": str(e)}

    try:
        ExcelService.save(wb, f"{result['action']}: {data.message[:60]} ({data.entered_by})")
    except Exception as e:
        raise HTTPException(500, f"Could not save the workbook to GitHub: {e}")

    return {"success": True, "type": "movement", **result}


@app.post("/api/undo")
def undo():
    wb = ExcelService.load_live_workbook()
    try:
        result = ExcelService.undo_last(wb)
    except ValueError as e:
        raise HTTPException(400, str(e))
    try:
        ExcelService.save(wb, "Undo last transaction")
    except Exception as e:
        raise HTTPException(500, f"Could not save the workbook to GitHub: {e}")
    return {"success": True, "reversed": result}


@app.get("/api/transactions")
def transactions(limit: int = 50):
    wb = ExcelService.load_live_workbook()
    return {"transactions": ExcelService.get_log_rows(wb, limit)}


@app.get("/api/products")
def products():
    """Current stock, grouped by product then location - powers the Reports table."""
    wb = ExcelService.load_live_workbook()
    summary = ExcelService.stock_summary(wb)
    out = []
    for product, locs in sorted(summary.items()):
        out.append({"product": product, "total": round(sum(locs.values()), 4),
                    "locations": [{"location": l, "quantity": q} for l, q in sorted(locs.items())]})
    return {"products": out}


@app.get("/api/aging-report")
def aging_report(threshold_days: Optional[int] = None):
    wb = ExcelService.load_live_workbook()
    rows = ExcelService.aging_batches(wb, threshold_days or AGING_THRESHOLD_DAYS)
    return {"threshold_days": threshold_days or AGING_THRESHOLD_DAYS, "batches": rows}


@app.get("/api/stock-by-location")
def stock_by_location():
    wb = ExcelService.load_live_workbook()
    summary = ExcelService.stock_summary(wb)
    totals: Dict[str, float] = {}
    for product, locs in summary.items():
        for loc, q in locs.items():
            totals[loc] = round(totals.get(loc, 0) + q, 4)
    return {"locations": [{"location": l, "quantity": q} for l, q in sorted(totals.items(), key=lambda x: -x[1])]}


@app.get("/api/dashboard-summary")
def dashboard_summary():
    wb = ExcelService.load_live_workbook()
    summary = ExcelService.stock_summary(wb)
    all_batches = [b for b in ExcelService.all_batches(wb) if (b["Qty Remaining"] or 0) > 0]
    aging = ExcelService.aging_batches(wb)
    total_qty = round(sum(b["Qty Remaining"] for b in all_batches), 4)
    return {
        "summary": {
            "total_products": len(summary),
            "total_locations": len(ExcelService.known_locations(wb)),
            "total_quantity": total_qty,
            "active_batches": len(all_batches),
            "aging_alerts": len(aging),
            "aging_threshold_days": AGING_THRESHOLD_DAYS,
        }
    }


@app.get("/api/download")
def download_excel():
    try:
        _, raw = _get_file_meta()
        if raw is None:
            wb = ExcelService.new_workbook()
            buf = io.BytesIO()
            wb.save(buf)
            raw = buf.getvalue()
    except Exception as e:
        raise HTTPException(500, f"Could not read the workbook: {e}")
    return StreamingResponse(
        io.BytesIO(raw),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventory_report.xlsx"},
    )


@app.get("/api/download-url")
def download_url_route():
    return {"url": get_download_url()}
